"""Write the 10-sheet xlsx workbook matching Almajed_FY2025_Financial_Data_updated.xlsx."""
from __future__ import annotations

from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .coa import CoA, load_coa
from .schemas import MasterRow, XLSX_HEADERS


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SUBTOTAL_FILL = PatternFill("solid", fgColor="DCE6F1")


def _readme_text(*, company: str, period_current: str, period_prior: str,
                 currency: str, approval_date: str, units_multiplier: int) -> list[list[str]]:
    units_label = {1: "actual", 1000: "thousands", 1_000_000: "millions"}.get(units_multiplier, str(units_multiplier))
    return [
        [f"{company} - {period_current} - Linked & Standardized Master Mapping"],
        [],
        [f"Period: {period_current} (with {period_prior} comparatives)"],
        [f"Currency: {currency}. Units: {units_label}. Approval: {approval_date or 'n/a'}"],
        [],
        ["PURPOSE"],
        ["Master mapping schema for multi-company comparison across Tadawul issuers."],
        ["Every numeric value links back to (a) the parent face-statement caption and"],
        ["(b) a standardized chart-of-accounts code so company-specific captions roll up"],
        ["to the same standard account across companies."],
        [],
        ["COLUMN SCHEMA (15 columns)"],
        ["  A  Parent Statement     - Balance Sheet / Income Statement / Cash Flow / Equity / Disclosure"],
        ["  B  Parent Section       - Parent grouping (e.g. Non-Current Assets)"],
        ["  C  Parent Caption       - Exact line item on the face of the parent statement"],
        ["  D  Std Parent Code      - Standardized code for the parent caption (cross-company key)"],
        ["  E  Std Parent Name      - Standardized name for the parent caption"],
        ["  F  Note Number          - Note reference (5, 6.1, 32(a), etc.)"],
        ["  G  Note Section         - Section within the note (NBV by Class, Movement, Detail, etc.)"],
        ["  H  Note Sub-Section     - Optional further sub-grouping"],
        ["  I  Line Item            - Company-specific caption as reported in the source"],
        ["  J  Std Item Code        - Standardized code for THIS row (cross-company key)"],
        ["  K  Std Item Name        - Standardized name for this row"],
        ["  L  Cross-Reference      - Other Std Codes this row also relates to"],
        ["  M  Row Type             - Primary | Note Detail | Disclosure"],
        [f"  N  {period_current}                 - {period_current} numeric value"],
        [f"  O  {period_prior}                 - {period_prior} numeric value (comparative)"],
        [],
        ["ROW TYPES"],
        ["  Primary      - Appears on the face of the BS/IS/CF/Equity statement"],
        ["  Note Detail  - Drilldown that ties back to a Primary caption"],
        ["  Disclosure   - No face anchor (capital commitments, contingencies, etc.)"],
        [],
        ["SHEETS"],
        ["  README                      - This sheet"],
        ["  Master Data                 - Every row, all statements + notes + disclosures"],
        ["  Standard CoA                - The standardized chart of accounts taxonomy"],
        ["  Balance Sheet (Linked)      - Filter: Parent Statement = Balance Sheet"],
        ["  Income Statement (Linked)   - Filter: Parent Statement = Income Statement"],
        ["  Cash Flow (Linked)          - Filter: Parent Statement = Cash Flow Statement"],
        ["  Equity (Linked)             - Filter: Parent Statement = Statement of Changes in Equity"],
        ["  Notes & Disclosures         - All Note Detail and Disclosure rows"],
        ["  Disclosures Only            - Only Row Type = Disclosure"],
        ["  Primary (Face) Only         - Only Row Type = Primary"],
        [],
        ["CONVENTIONS"],
        ["  - Negatives shown as negative numbers; zeros as 0"],
        ["  - IS and CF expenses are negative; BS items in natural sign"],
        [f"  - All values consolidated Group level, in {currency}"],
    ]


def _write_data_sheet(ws, rows: Iterable[MasterRow], period_current: str,
                      period_prior: str) -> None:
    headers = XLSX_HEADERS + [period_current, period_prior]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r in rows:
        ws.append(r.as_xlsx_row())
    # Reasonable column widths
    widths = [22, 22, 36, 14, 28, 10, 18, 16, 36, 14, 28, 18, 12, 16, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def _filter(rows, **kw):
    out = []
    for r in rows:
        ok = True
        for k, v in kw.items():
            if isinstance(v, (set, tuple, list)):
                if getattr(r, k) not in v:
                    ok = False; break
            else:
                if getattr(r, k) != v:
                    ok = False; break
        if ok:
            out.append(r)
    return out


def write_xlsx(out_path: Path | str, *, rows: list[MasterRow],
               company: str, period_current: str, period_prior: str,
               currency: str, approval_date: str, units_multiplier: int,
               coa: CoA | None = None) -> None:
    coa = coa or load_coa()
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # README
    ws = wb.active
    ws.title = "README"
    for line in _readme_text(company=company, period_current=period_current,
                              period_prior=period_prior, currency=currency,
                              approval_date=approval_date,
                              units_multiplier=units_multiplier):
        ws.append(line)
    ws["A1"].font = Font(bold=True, size=14)
    ws.column_dimensions["A"].width = 110

    # Master Data
    _write_data_sheet(wb.create_sheet("Master Data"), rows, period_current, period_prior)

    # Standard CoA
    coa_ws = wb.create_sheet("Standard CoA")
    coa_ws.append(["Std Code", "Std Name", "Std Category", "Statement"])
    for c in coa_ws[1]:
        c.fill = HEADER_FILL; c.font = HEADER_FONT
    for a in coa.accounts:
        coa_ws.append([a.code, a.name, a.category, a.statement])
    for i, w in enumerate([14, 40, 22, 28], start=1):
        coa_ws.column_dimensions[get_column_letter(i)].width = w
    coa_ws.freeze_panes = "A2"

    # Per-statement filtered views.
    for sheet_name, stmt in [
        ("Balance Sheet (Linked)", "Balance Sheet"),
        ("Income Statement (Linked)", "Income Statement"),
        ("Cash Flow (Linked)", "Cash Flow Statement"),
        ("Equity (Linked)", "Statement of Changes in Equity"),
    ]:
        _write_data_sheet(
            wb.create_sheet(sheet_name),
            _filter(rows, parent_statement=stmt),
            period_current, period_prior,
        )

    # Notes & Disclosures
    _write_data_sheet(
        wb.create_sheet("Notes & Disclosures"),
        _filter(rows, row_type={"Note Detail", "Disclosure"}),
        period_current, period_prior,
    )
    # Disclosures Only
    _write_data_sheet(
        wb.create_sheet("Disclosures Only"),
        _filter(rows, row_type="Disclosure"),
        period_current, period_prior,
    )
    # Primary (Face) Only
    _write_data_sheet(
        wb.create_sheet("Primary (Face) Only"),
        _filter(rows, row_type="Primary"),
        period_current, period_prior,
    )

    wb.save(out_path)
